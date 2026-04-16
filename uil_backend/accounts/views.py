from django.contrib.auth import get_user_model
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.http import HttpResponse
from django.db import  transaction
from django.conf import settings
from django_filters.rest_framework import DjangoFilterBackend
from django.shortcuts import get_object_or_404

from datetime import timedelta

from rest_framework import status as drf_status
from rest_framework_simplejwt.exceptions import TokenError, InvalidToken
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.filters import OrderingFilter,SearchFilter
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.exceptions import ValidationError, NotFound, AuthenticationFailed, PermissionDenied,NotAuthenticated
from rest_framework.permissions import  AllowAny
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status as http_status
from rest_framework.viewsets import ReadOnlyModelViewSet
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import ValidationError, NotFound, PermissionDenied, NotAuthenticated


from djoser.views import UserViewSet
from djoser.compat import get_user_email
from djoser import signals

from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from authorization.permissions import HasRequiredPermissions,IsOwnerOrHasRequiredPermissions
from authorization.utilis import get_scope
from authorization.decorators import require_permissions
from .services import read_users_from_excel
from .emails import  CustomConfirmationEmail, CustomOneTimePasswordEmail, CustomRejectionEmail
import logging
from .models import User
from organizational_structure.models import OrganizationalUnit
from .serializers import (
    BulkUserSerializer,
    FirstTimeChangePasswordSerializer,
    DashboardSummarySerializer,
    UploadUsersExcelSerializer,
    UserSerializer,
    CustomTokenObtainPairSerializer,
    UserFullSerializer
)
from .paginations import UserPagination
from .permissions import USER_REQUIRED_PERMISSIONS

User = get_user_model()
logger = logging.getLogger(__name__)

class PrivateUserRetrieveView(APIView):
    def get(self, request):
        token = request.headers.get("X-API-KEY")

        if token != settings.PRIVATE_API_TOKEN:
            print(token,settings.PRIVATE_API_TOKEN)
            return Response({"detail": "Unauthorized"}, status=403)

        email = request.query_params.get("email")
        if not email:
            return Response({"detail": "email is required"}, status=400)

        user = get_object_or_404(
            User.objects.select_related("researcherprofile"),
            email=email
        )

        serializer = UserFullSerializer(user)
        return Response(serializer.data)
class CustomUserViewSet(UserViewSet):
    pagination_class = UserPagination
    filter_backends = [DjangoFilterBackend,OrderingFilter,SearchFilter]
    filterset_fields = ['status','is_superuser']
    ordering_fields = ['created_at', 'updated_at','first_name']
    search_fields = ['username', 'first_name', 'father_name', 'grand_father_name','email']
    target_scope=None
    

    def get_permissions(self):
        """setting permission according to the  action and also adding permission class depending on action"""
        self.required_permissions = USER_REQUIRED_PERMISSIONS.get(self.action, [])
        if self.action in( "create","reset_password","reset_password_confirm"):
            permission_classes = [AllowAny]
            
        elif self.action in ("update", "partial_update","destroy"):
            permission_classes = [IsAuthenticated, IsOwnerOrHasRequiredPermissions]

        elif self.action == "upload_users_excel":
            academic_unit_id = self.request.data.get("academic_unit")
            if academic_unit_id:
                try:
                    self.target_scope = OrganizationalUnit.objects.get(id=int(academic_unit_id))
                except OrganizationalUnit.DoesNotExist:
                    self.target_scope = None
            else:
                self.target_scope = None
            permission_classes = [HasRequiredPermissions]
        else:
            permission_classes = [HasRequiredPermissions]
        return [permission() for permission in permission_classes]

    def get_object(self):
        """it pass the scope of the target to the class and check object permission"""
        obj = super().get_object()
        self.target_scope = obj.academic_unit
        self.check_object_permissions(self.request, obj)
        return obj

    def get_queryset(self):
        """enbales the list to be returned by the  scope of administatror"""
        user = self.request.user
        if not user.is_authenticated:
            raise NotAuthenticated("Authentication credentials were not provided")

        scope = get_scope(user, self.required_permissions)
        parent_unit_id = self.request.query_params.get("academic_unit_scope")
        if parent_unit_id:
            try:
                parent_unit = OrganizationalUnit.objects.get(
                    id=int(parent_unit_id))
            except OrganizationalUnit.DoesNotExist:
                return User.objects.none()
            if not parent_unit in scope:
                return User.objects.none()
            self.request.parent_scope = parent_unit

            filter_scope = [parent_unit.id] + \
                [u.id for u in parent_unit.get_all_descendants()]
            scope_qs = OrganizationalUnit.objects.filter(id__in=filter_scope)
        else:
            scope_qs = scope



        queryset = User.objects.filter(
            academic_unit__in=scope_qs
        ).order_by('-created_at')

        has_profile_param = self.request.query_params.get('has_profile')
        if has_profile_param is not None:
            if has_profile_param.lower() in ['true', '1']:
                queryset = queryset.filter(researcherprofile__isnull=False)
            else:
                queryset = queryset.filter(researcherprofile__isnull=True)
        return queryset

    def perform_create(self, serializer, *args, **kwargs):
        """sending email while creating user """
        user = serializer.save()
        user.created_by = user
        user.updated_by = user
        user.status='APPROVED'
        user.save(update_fields=['created_by', 'updated_by', 'status'])
        signals.user_registered.send(
            sender=self.__class__, user=user, request=self.request)


        # protocol = self.request.scheme
        # domain = self.request.get_host()
        # uid = urlsafe_base64_encode(str(user.pk).encode())
        # token = default_token_generator.make_token(user)
        # email = AdminActivationEmail(
        #     user=user,
        #     protocol=protocol,
        #     domain=domain,
        #     uid=uid,
        #     token=token
        # )
        # email.send()
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.is_superuser:
            return Response(
                {"detail": "Cannot delete a superuser."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)
    @action(detail=False,methods=["post"])
    def first_time_change_password(self, request):
        serializer = FirstTimeChangePasswordSerializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)

        user = request.user
        user.set_password(serializer.validated_data["new_password"])
        user.must_change_password = False
        user.raw_password = ""
        user.save(update_fields=["password","must_change_password", "raw_password"])

        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["post"], url_path="upload-users")
    def upload_users_excel(self, request):
        serializer = UploadUsersExcelSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        excel_file = serializer.validated_data["file"]
        academic_unit = serializer.validated_data["academic_unit"]
        self.target_scope = academic_unit.id

        users_data = read_users_from_excel(excel_file)
        BulkUserSerializer.validate_bulk(users_data)


        errors = []
        created_users = []
        for idx, user_data in enumerate(users_data, start=1):
            username = user_data.get("username")
            email = user_data.get("email")
            if User.objects.filter(username=username).exists():
                errors.append({
                    "row": idx,
                    "field": "username",
                    "message": f"Username '{username}' is already registered."
                })
                continue
            if User.objects.filter(email=email).exists():
                errors.append({
                    "row": idx,
                    "field": "email",
                    "message": f"Email '{email}' is already registered."
                })
                continue

            serializer = BulkUserSerializer(data=user_data)
            if serializer.is_valid():
                try:
                    with transaction.atomic():
                        user = serializer.save()
                        otp_password = get_random_string(length=8)
                        user.set_password(otp_password)
                        user.raw_password = otp_password
                        user.must_change_password = True
                        user.created_by=self.request.user
                        user.academic_unit=academic_unit
                        user.save()
                        email_obj = CustomOneTimePasswordEmail(user, otp_password)
                        email_obj.send([user.email])
                        created_users.append(user)
                except Exception as e:
                    errors.append({
                        "row": idx,
                        "message": f"Unexpected error saving user: {str(e)}"
                    })
            else:
                errors.append({
                    "row": idx,
                    "message": serializer.errors
                })
        if errors:
            return Response({"created": len(created_users), "errors": errors},
                            status=status.HTTP_400_BAD_REQUEST)
        return Response({"created": len(created_users)}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"], url_path="resend-credentials")
    def resend_credentials(self, request):
        user_id = request.data.get("user_id")
        user_ids = request.data.get("user_ids")
        if not user_id and not user_ids:
            raise ValidationError({"detail": "Provide 'user_id' or 'user_ids'."})
        if user_id:
            user_ids = [user_id]
        elif not isinstance(user_ids, list):
            raise ValidationError({"user_ids": "Must be a list of user IDs."})
        result = {"sent": [], "skipped": {}, "not_found": []}
        for uid in user_ids:
            try:
                user = User.objects.get(id=uid)
                if not user.raw_password:
                    result["skipped"][uid] = "Password unavailable"
                    continue
                email_obj = CustomOneTimePasswordEmail(user, user.raw_password)
                email_obj.send([user.email])
                result["sent"].append(uid)
            except User.DoesNotExist:
                result["not_found"].append(uid)
            except Exception as e:
                result["skipped"][uid] = "Unexpected error"
                logger.exception(f"Failed to create user {uid}: {e}")

        return Response(result, status=status.HTTP_200_OK)

    @action(['PATCH'] ,detail=True)
    def patch_status(self,request, id=None):
        try:
            user = self.get_object()
            new_status = request.data.get("status")
            if new_status not in ["APPROVED", "PENDING", "REJECTED"]:
                raise ValidationError({"status": "Invalid status value."})

            user.status = new_status
            user.save()

            if new_status == "APPROVED":
                context = {"user": user}
                to = [get_user_email(user)]
                CustomConfirmationEmail(request, context).send(to)
            elif new_status == "REJECTED":
                context = {"user": user}
                to = [get_user_email(user)]
                CustomRejectionEmail(request, context).send(to)

            return Response({"id": user.id, "status": new_status}, status=drf_status.HTTP_200_OK)

        except User.DoesNotExist:
            return Response({"error": "User not found."}, status=drf_status.HTTP_404_NOT_FOUND)

    def activation(self, request, *args, **kwargs):
        return Response({"detail": "Method not allowed."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

    def resend_activation(self, request, *args, **kwargs):
        return Response({"detail": "Method not allowed."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

    def set_username(self, request, *args, **kwargs):
        return Response({"detail": "Method not allowed."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

    def reset_username(self, request, *args, **kwargs):
        return Response({"detail": "Method not allowed."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

    def reset_username_confirm(self, request, *args, **kwargs):
        return Response({"detail": "Method not allowed."}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

    # @action(["GET"],detail=False, url_path="activate-user")
    # def activate_user_from_email(self, request, uid, token):
    #     try:
    #         uid = urlsafe_base64_decode(uid).decode()
    #         user = User.objects.get(pk=uid)
            # self.check_object_permissions(request, user)
            # check_object_permission(
            #     request=request,
            #     permission_class=HasRequiredPermissions,
            #     obj=user,
            #     required_permissions=USER_REQUIRED_PERMISSIONS.get('patch_status', []),
            #     target_scope=user.academic_unit
            # )

        #     if not default_token_generator.check_token(user, token):
        #         raise ValidationError("Invalid or expired activation link.")
        #     context = {"user": user}
        #     to = [get_user_email(user)]
        #     user.status = 'APPROVED'
        #     user.save()
        #     CustomConfirmationEmail(request, context).send(to)
        #     return Response({"detail": "User has been activated successfully."}, status=status.HTTP_200_OK)
        # except User.DoesNotExist:
        #     raise NotFound("User not found.")
        # except ValueError:
        #     raise ValidationError("Invalid UID format.")


class PublicUserViewSet(ReadOnlyModelViewSet):
    pagination_class = UserPagination
    permission_classes = [AllowAny]
    serializer_class=UserSerializer
    queryset = User.objects.all().select_related(
        'academic_unit'
    ).order_by('-created_at')

    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_fields = ['status', 'is_superuser']
    ordering_fields = ['created_at', 'updated_at', 'first_name']
    search_fields = ['username', 'first_name', 'father_name', 'grand_father_name', 'email']




@require_permissions(USER_REQUIRED_PERMISSIONS.get("dashboard_summary"))
@api_view(['GET'])
@permission_classes([HasRequiredPermissions])
def dashboard_summary(request):
    scope = get_scope(request.user,USER_REQUIRED_PERMISSIONS.get("dashboard_summary"))
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    total_users = User.objects.filter(            academic_unit__in=scope
).count()
    pending_users = User.objects.filter(status='PENDING',            academic_unit__in=scope
).count()
    approved_users = User.objects.filter(status='APPROVED',            academic_unit__in=scope
).count()
    rejected_users = User.objects.filter(status='REJECTED',            academic_unit__in=scope
).count()
    active_users = User.objects.filter(is_active=True,            academic_unit__in=scope
).count()
    inactive_users = User.objects.filter(is_active=False,            academic_unit__in=scope
).count()
    superusers = User.objects.filter(is_superuser=True,            academic_unit__in=scope
).count()
    users_created_today = User.objects.filter(created_at__date=today,            academic_unit__in=scope
).count()
    users_created_this_week = User.objects.filter(
                    academic_unit__in=scope,

        created_at__date__gte=week_ago).count()
    users_created_this_month = User.objects.filter(
                    academic_unit__in=scope,

        created_at__date__gte=month_ago).count()

    summary_data = {
        'total_users': total_users,
        'pending_users': pending_users,
        'approved_users': approved_users,
        'rejected_users': rejected_users,
        'active_users': active_users,
        'inactive_users': inactive_users,
        'superusers': superusers,
        'users_created_today': users_created_today,
        'users_created_this_week': users_created_this_week,
        'users_created_this_month': users_created_this_month,
    }

    serializer = DashboardSummarySerializer(summary_data)
    return Response(serializer.data)



# @require_permissions(USER_REQUIRED_PERMISSIONS.get("user_excel_by_academic_unit"))
@api_view(["GET"])
# @permission_classes([HasRequiredPermissions])
def user_excel_by_academic_unit(request, pk):
    """
    Generate Excel file of researchers grouped by direct child academic units
    under the given parent OrganizationalUnit (pk).
    Includes all descendants (grandchildren) automatically.
    """
    parent_unit = OrganizationalUnit.objects.filter(id=pk).first()
    if not parent_unit:
        return Response(
            {"error": f"Academic unit with id {pk} not found."},
            status=http_status.HTTP_404_NOT_FOUND
        )
    child_units = parent_unit.get_children()

    if not child_units.exists():
        child_units = [parent_unit]

    # 3️⃣ Helper to auto-adjust Excel columns
    def auto_adjust(ws, min_width=3, max_width=40):
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(min_width, min(max_length + 2, max_width))

    # 4️⃣ Create a workbook
    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.append(["No", "Academic Unit", "No. of Researchers"])
    total = 0

    for idx, unit in enumerate(child_units, start=1):
        descendants = [unit] + unit.get_all_descendants()
        descendant_ids = [u.id for u in descendants]

        user = User.objects.filter(
            academic_unit__id__in=descendant_ids
        ).select_related( "academic_unit")

        if not user.exists():
            continue

        # Create sheet per child
        sheet_name = (unit.abbreviation or unit.name)[:31]
        ws_fac = wb.create_sheet(title=sheet_name)

        headers = [
            "No",
            "Full Name",
            "Email",
            "Academic Unit",
        ]
        ws_fac.append(headers)

        for n, r in enumerate(user, start=1):
            ws_fac.append([
                n,
                r.get_full_name(),
                r.email,
                r.academic_unit.name if r.academic_unit else "",
            ])

        auto_adjust(ws_fac)

        ws_summary.append([idx, unit.name, user.count()])
        total += user.count()

    ws_summary.append(["", "Total", total])
    last_row = ws_summary.max_row
    ws_summary[f"B{last_row}"].font = Font(bold=True)
    ws_summary[f"C{last_row}"].font = Font(bold=True)
    auto_adjust(ws_summary)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Researchers_Under_{parent_unit.name}.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
    def post(self, request, *args, **kwargs):
        try:
            response = super().post(request, *args, **kwargs)
        except AuthenticationFailed:
            raise AuthenticationFailed("Incorrect username or password.")

        refresh_token = response.data.get("refresh")
        if not refresh_token:
            raise AuthenticationFailed("No refresh token returned.")

        # Set refresh token in httpOnly cookie
        response.set_cookie(
            key="refresh_token",
            value=refresh_token,
            httponly=True,
            secure=True,  # True if using HTTPS
            samesite="Strict",  # or "Lax" depending on your use-case
            max_age=7*24*60*60  # 7 days in seconds
        )

        # Optionally remove refresh token from response body
        response.data.pop("refresh", None)

        # Validate user status
        try:
            token = RefreshToken(refresh_token)
            user_id = token.get("user_id")
            user = get_user_model().objects.get(id=user_id)
        except (TokenError, InvalidToken):
            raise AuthenticationFailed("Session expired. Please login again.")
        except get_user_model().DoesNotExist:
            raise NotFound("User not found.")

        if user.status == 'PENDING':
            raise PermissionDenied("Your account is pending approval.")
        elif user.status == 'REJECTED':
            raise PermissionDenied("Your account has been rejected.")

        return response